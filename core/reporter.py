"""
Reporter stage – generates multi-format comparison reports.

Produces professional output in Excel (.xlsx), CSV, HTML, and Markdown
formats.  Each format includes configurable sections: Summary, Mismatch
Details, Unmatched Records, Per-Column Breakdown, Normalization Log,
and Filter Log.

Rule catalogue
~~~~~~~~~~~~~~
RPT-001  Output directory is created automatically if it does not exist.
RPT-002  All configured formats are generated independently; a failure
          in one format does not block others.
RPT-003  Empty sections produce a clear "no data" message instead of
          being silently omitted.
RPT-004  Excel workbooks use conditional formatting and frozen headers.
RPT-005  HTML output is self-contained with inline CSS.
RPT-006  File names are resolved relative to the output directory.

Configuration (top-level key ``report``)
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
.. code-block:: yaml

    report:
      title: "Data Comparison Report"
      output:
        formats:
          - type: "excel"
            file_name: "comparison_report.xlsx"
          - type: "html"
            file_name: "comparison_report.html"
          - type: "csv"
            file_name: "comparison_report"
          - type: "markdown"
            file_name: "comparison_report.md"
      sections:
        - name: "summary"
          enabled: true
        - name: "mismatch_details"
          enabled: true
        - name: "unmatched_records"
          enabled: true
        - name: "per_column_breakdown"
          enabled: true
        - name: "normalization_log"
          enabled: true
        - name: "filter_log"
          enabled: true
"""
from __future__ import annotations

import csv
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from .engine import (
    ConfigurationError,
    OutputError,
    PipelineContext,
    PipelineStage,
)

# openpyxl is an optional dependency – import styles at module level
# so that static helper methods can reference them without re-importing.
try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False
    Font = None  # type: ignore[assignment, misc]
    PatternFill = None  # type: ignore[assignment, misc]
    Alignment = None  # type: ignore[assignment, misc]
    Border = None  # type: ignore[assignment, misc]
    Side = None  # type: ignore[assignment, misc]


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Conditional formatting colours
_MATCH_FILL_HEX: str = "C6EFCE"   # green
_MISMATCH_FILL_HEX: str = "FFC7CE"  # red
_HEADER_FILL_HEX: str = "4472C4"    # accent blue
_HEADER_FONT_COLOR: str = "FFFFFF"

# Default section names
_DEFAULT_SECTIONS: List[str] = [
    "summary",
    "mismatch_details",
    "unmatched_records",
    "per_column_breakdown",
    "normalization_log",
    "filter_log",
]


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def _html_escape(text: str) -> str:
    """Escape HTML special characters."""
    return (
        text
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#39;")
    )


# ---------------------------------------------------------------------------
# ReporterStage
# ---------------------------------------------------------------------------

class ReporterStage(PipelineStage):
    """Pipeline stage that generates comparison reports in multiple formats.

    Supports Excel (.xlsx), CSV, HTML, and Markdown output.  Each format
    includes all enabled sections from the report configuration.

    See the module docstring for the configuration reference.
    """

    name: str = "report"

    # ------------------------------------------------------------------
    # Configuration validation
    # ------------------------------------------------------------------

    def validate_config(self, config_section: Dict) -> None:
        """Validate the ``report`` section of the pipeline configuration.

        Parameters
        ----------
        config_section:
            The value of ``config.get("report", {})``.

        Raises
        ------
        ConfigurationError
            On any structural or semantic issue.
        """
        if config_section is None:
            return

        if not isinstance(config_section, dict):
            raise ConfigurationError(
                "'report' configuration must be a mapping.",
                {"rule": "RPT-001", "got": type(config_section).__name__},
            )

        # Validate output.formats
        output_cfg = config_section.get("output")
        if output_cfg is not None:
            if not isinstance(output_cfg, dict):
                raise ConfigurationError(
                    "'report.output' must be a mapping.",
                    {"rule": "RPT-001", "got": type(output_cfg).__name__},
                )

            formats = output_cfg.get("formats", [])
            if not isinstance(formats, list):
                raise ConfigurationError(
                    "'report.output.formats' must be a list.",
                    {"rule": "RPT-001", "got": type(formats).__name__},
                )

            valid_types = {"excel", "html", "csv", "markdown"}
            for idx, fmt in enumerate(formats):
                if not isinstance(fmt, dict):
                    raise ConfigurationError(
                        f"Format entry #{idx} must be a mapping.",
                        {"rule": "RPT-001", "index": idx},
                    )
                fmt_type = fmt.get("type", "")
                if fmt_type not in valid_types:
                    raise ConfigurationError(
                        f"Format entry #{idx}: unknown type '{fmt_type}'. "
                        f"Valid types: {sorted(valid_types)}.",
                        {"rule": "RPT-001", "index": idx, "type": fmt_type},
                    )

        # Validate sections
        sections = config_section.get("sections", [])
        if sections is not None:
            if not isinstance(sections, list):
                raise ConfigurationError(
                    "'report.sections' must be a list.",
                    {"rule": "RPT-001", "got": type(sections).__name__},
                )
            for idx, section in enumerate(sections):
                if not isinstance(section, dict):
                    raise ConfigurationError(
                        f"Section entry #{idx} must be a mapping.",
                        {"rule": "RPT-001", "index": idx},
                    )
                if "name" not in section:
                    raise ConfigurationError(
                        f"Section entry #{idx} is missing 'name'.",
                        {"rule": "RPT-001", "index": idx},
                    )

    # ------------------------------------------------------------------
    # Execution
    # ------------------------------------------------------------------

    def execute(self, context: PipelineContext, config: Dict) -> None:
        """Generate all configured output formats.

        Each format is generated independently; a failure in one does
        not block others (RPT-002).

        Parameters
        ----------
        context:
            The shared pipeline context (must contain comparison results).
        config:
            The full top-level pipeline configuration dictionary.
        """
        from time import perf_counter

        t0 = perf_counter()

        report_cfg = config.get("report", {})
        title = report_cfg.get("title", "Data Comparison Report")
        output_dir = context.output_dir

        # RPT-001: ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)

        # Determine enabled sections
        enabled_sections = self._get_enabled_sections(report_cfg)

        # Determine configured formats
        output_formats = report_cfg.get("output", {}).get("formats", [])
        if not output_formats:
            # Default: generate HTML
            output_formats = [{"type": "html", "file_name": "comparison_report.html"}]

        results: Dict[str, Any] = {}
        errors: List[str] = []

        for fmt_cfg in output_formats:
            fmt_type = fmt_cfg.get("type", "")
            file_name = fmt_cfg.get("file_name", "")

            try:
                if fmt_type == "excel":
                    output_path = os.path.join(output_dir, file_name)
                    self._generate_excel(context, config, output_path, enabled_sections, title)
                    results[fmt_type] = output_path

                elif fmt_type == "csv":
                    base_name = file_name or "comparison_report"
                    csv_dir = os.path.join(output_dir, base_name)
                    self._generate_csv(context, config, csv_dir, enabled_sections)
                    results[fmt_type] = csv_dir

                elif fmt_type == "html":
                    output_path = os.path.join(output_dir, file_name)
                    self._generate_html(context, config, output_path, enabled_sections, title)
                    results[fmt_type] = output_path

                elif fmt_type == "markdown":
                    output_path = os.path.join(output_dir, file_name)
                    self._generate_markdown(context, config, output_path, enabled_sections, title)
                    results[fmt_type] = output_path

            except OutputError:
                errors.append(f"{fmt_type}: unknown output error")
            except Exception as exc:
                errors.append(f"{fmt_type}: {exc}")

        elapsed = perf_counter() - t0
        context.stats["report"] = {
            "formats_generated": list(results.keys()),
            "formats_failed": errors,
            "sections_enabled": enabled_sections,
            "elapsed_seconds": round(elapsed, 4),
        }

        for fmt_type, path in results.items():
            context.stats["report"][f"{fmt_type}_path"] = path

    # ------------------------------------------------------------------
    # Section routing
    # ------------------------------------------------------------------

    def _get_enabled_sections(self, report_cfg: Dict) -> List[str]:
        """Read report config to determine which sections to generate.

        Returns
        -------
        list[str]
            Names of enabled sections.
        """
        sections = report_cfg.get("sections", [])

        if not sections:
            return list(_DEFAULT_SECTIONS)

        enabled: List[str] = []
        for section in sections:
            if isinstance(section, dict) and section.get("enabled", True):
                name = section.get("name", "")
                if name:
                    enabled.append(name)

        return enabled

    # ------------------------------------------------------------------
    # Data extraction helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _get_summary_data(context: PipelineContext, stats: Dict) -> Dict[str, Any]:
        """Build the summary section data from context and stats."""
        comparison = context.comparison_result or {}
        matched_keys = comparison.get("matched_keys", [])
        only_in_a = comparison.get("only_in_a", [])
        only_in_b = comparison.get("only_in_b", [])
        details = comparison.get("details", [])

        mismatches = [d for d in details if d.get("status") == "mismatch"]
        matches = [d for d in details if d.get("status") == "match"]

        # Per-key full match detection
        key_statuses: Dict[str, bool] = {}
        for d in details:
            key = d.get("key", "")
            if key not in key_statuses:
                key_statuses[key] = True
            if d.get("status") == "mismatch":
                key_statuses[key] = False

        full_match_count = sum(1 for v in key_statuses.values() if v)
        total_compared = len(key_statuses)
        match_rate = round((full_match_count / total_compared * 100) if total_compared > 0 else 100.0, 2)

        return {
            "title": "Summary",
            "total_rows_source_a": context.rows_source_a,
            "total_rows_source_b": context.rows_source_b,
            "total_compared_keys": total_compared,
            "matched_keys_count": len(matched_keys),
            "only_in_a_count": len(only_in_a),
            "only_in_b_count": len(only_in_b),
            "full_match_count": full_match_count,
            "partial_mismatch_count": total_compared - full_match_count,
            "total_cell_mismatches": len(mismatches),
            "total_cell_matches": len(matches),
            "match_rate_percent": match_rate,
            "only_in_a": only_in_a[:100],
            "only_in_b": only_in_b[:100],
        }

    @staticmethod
    def _get_mismatch_details(context: PipelineContext) -> List[Dict[str, Any]]:
        """Extract mismatch detail rows."""
        comparison = context.comparison_result or {}
        details = comparison.get("details", [])
        return [d for d in details if d.get("status") == "mismatch"]

    @staticmethod
    def _get_unmatched_records(context: PipelineContext) -> Dict[str, List]:
        """Extract unmatched record keys."""
        comparison = context.comparison_result or {}
        return {
            "only_in_a": comparison.get("only_in_a", []),
            "only_in_b": comparison.get("only_in_b", []),
        }

    @staticmethod
    def _get_per_column_breakdown(context: PipelineContext) -> List[Dict[str, Any]]:
        """Build per-column match/mismatch breakdown."""
        comparison = context.comparison_result or {}
        details = comparison.get("details", [])

        column_data: Dict[str, Dict[str, int]] = {}
        for d in details:
            col = d.get("column", "")
            if col not in column_data:
                column_data[col] = {"matched": 0, "mismatched": 0}
            if d.get("status") == "match":
                column_data[col]["matched"] += 1
            else:
                column_data[col]["mismatched"] += 1

        rows: List[Dict[str, Any]] = []
        for col, counts in sorted(column_data.items()):
            total = counts["matched"] + counts["mismatched"]
            rate = round((counts["matched"] / total * 100) if total > 0 else 100.0, 2)
            rows.append({
                "column": col,
                "matched": counts["matched"],
                "mismatched": counts["mismatched"],
                "total_compared": total,
                "match_rate_percent": rate,
            })
        return rows

    # ------------------------------------------------------------------
    # Excel generation
    # ------------------------------------------------------------------

    def _generate_excel(
        self,
        context: PipelineContext,
        config: Dict,
        output_path: str,
        sections: List[str],
        title: str,
    ) -> None:
        """Generate a multi-sheet Excel workbook with openpyxl.

        Parameters
        ----------
        context:
            Pipeline context with comparison results.
        config:
            Full pipeline configuration.
        output_path:
            File path for the .xlsx output.
        sections:
            List of enabled section names.
        title:
            Report title (used in the Summary sheet header).
        """
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise OutputError(
                "openpyxl is required for Excel output. "
                "Install it with: pip install openpyxl",
                {"rule": "RPT-001", "pip_package": "openpyxl"},
            ) from exc

        if not _HAS_OPENPYXL:
            raise OutputError(
                "openpyxl is required for Excel output. "
                "Install it with: pip install openpyxl",
                {"rule": "RPT-001", "pip_package": "openpyxl"},
            )

        wb = Workbook()
        stats = context.stats or {}

        # Styles
        header_font = Font(name="Calibri", bold=True, color=_HEADER_FONT_COLOR, size=11)
        header_fill = PatternFill(start_color=_HEADER_FILL_HEX, end_color=_HEADER_FILL_HEX, fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        match_fill = PatternFill(start_color=_MATCH_FILL_HEX, end_color=_MATCH_FILL_HEX, fill_type="solid")
        mismatch_fill = PatternFill(start_color=_MISMATCH_FILL_HEX, end_color=_MISMATCH_FILL_HEX, fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        first_sheet = True

        for section_name in sections:
            ws = wb.active if first_sheet else wb.create_sheet(title=section_name.replace("_", " ").title())
            first_sheet = False

            content = self._apply_section_content(section_name, context, config)

            if section_name == "summary":
                ws.title = "Summary"
                self._write_summary_sheet(ws, content, title, header_font, header_fill, header_alignment, thin_border)
            elif section_name == "mismatch_details":
                ws.title = "Mismatch Details"
                self._write_table_sheet(
                    ws, content,
                    headers=["Key", "Column", "Value A", "Value B", "Diff Type"],
                    row_keys=["key", "column", "value_a", "value_b", "diff_type"],
                    header_font=header_font, header_fill=header_fill,
                    header_alignment=header_alignment, thin_border=thin_border,
                    highlight_fill=mismatch_fill,
                )
            elif section_name == "unmatched_records":
                ws.title = "Unmatched Records"
                self._write_unmatched_sheet(ws, content, header_font, header_fill, header_alignment, thin_border)
            elif section_name == "per_column_breakdown":
                ws.title = "Per-Column Breakdown"
                self._write_table_sheet(
                    ws, content,
                    headers=["Column", "Matched", "Mismatched", "Total", "Match Rate %"],
                    row_keys=["column", "matched", "mismatched", "total_compared", "match_rate_percent"],
                    header_font=header_font, header_fill=header_fill,
                    header_alignment=header_alignment, thin_border=thin_border,
                )
            elif section_name == "normalization_log":
                ws.title = "Normalization Log"
                self._write_log_sheet(ws, context.normalization_log, header_font, header_fill, header_alignment, thin_border)
            elif section_name == "filter_log":
                ws.title = "Filter Log"
                self._write_log_sheet(ws, context.filter_log, header_font, header_fill, header_alignment, thin_border)

            # Auto-adjust column widths
            for col_cells in ws.columns:
                max_length = 0
                first_real = None
                for cell in col_cells:
                    # Skip MergedCell objects which lack column_letter
                    if not hasattr(cell, "column_letter"):
                        continue
                    if first_real is None:
                        first_real = cell
                    try:
                        val = str(cell.value) if cell.value is not None else ""
                        if len(val) > max_length:
                            max_length = len(val)
                    except Exception:
                        pass
                if first_real is not None:
                    col_letter = first_real.column_letter
                    adjusted = min(max_length + 2, 50)
                    ws.column_dimensions[col_letter].width = max(adjusted, 12)

        # Remove default empty sheet if we created sheets manually
        if len(wb.sheetnames) > 1 and wb.sheetnames[0] == "Sheet":
            del wb["Sheet"]

        try:
            wb.save(output_path)
        except Exception as exc:
            raise OutputError(
                f"Failed to write Excel file: {output_path}",
                {"rule": "RPT-002", "path": output_path, "error": str(exc)},
            ) from exc

    @staticmethod
    def _write_summary_sheet(
        ws: Any,
        summary: Dict[str, Any],
        title: str,
        header_font: Any,
        header_fill: Any,
        header_alignment: Any,
        thin_border: Any,
    ) -> None:
        """Write the Summary sheet."""
        # Title
        ws.merge_cells("A1:B1")
        title_cell = ws["A1"]
        title_cell.value = title
        title_cell.font = Font(name="Calibri", bold=True, size=14, color=_HEADER_FONT_COLOR)

        ws.merge_cells("A2:B2")
        timestamp_cell = ws["A2"]
        timestamp_cell.value = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"
        timestamp_cell.font = Font(name="Calibri", italic=True, size=10, color="666666")

        # Summary metrics
        row_idx = 4
        metrics = [
            ("Total Rows (Source A)", summary.get("total_rows_source_a", 0)),
            ("Total Rows (Source B)", summary.get("total_rows_source_b", 0)),
            ("Keys Compared", summary.get("total_compared_keys", 0)),
            ("Fully Matched Keys", summary.get("full_match_count", 0)),
            ("Keys with Mismatches", summary.get("partial_mismatch_count", 0)),
            ("Only in Source A", summary.get("only_in_a_count", 0)),
            ("Only in Source B", summary.get("only_in_b_count", 0)),
            ("Cell-Level Matches", summary.get("total_cell_matches", 0)),
            ("Cell-Level Mismatches", summary.get("total_cell_mismatches", 0)),
            ("Match Rate (%)", summary.get("match_rate_percent", 0)),
        ]

        # Header row
        for col_idx, header_text in enumerate(["Metric", "Value"], start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        row_idx += 1

        for metric_name, metric_value in metrics:
            cell_a = ws.cell(row=row_idx, column=1, value=metric_name)
            cell_a.border = thin_border
            cell_b = ws.cell(row=row_idx, column=2, value=metric_value)
            cell_b.border = thin_border
            cell_b.alignment = Alignment(horizontal="right")

            # Highlight match rate
            if "Match Rate" in metric_name:
                try:
                    rate = float(metric_value)
                    if rate >= 90:
                        cell_b.fill = PatternFill(
                            start_color=_MATCH_FILL_HEX,
                            end_color=_MATCH_FILL_HEX,
                            fill_type="solid",
                        )
                    elif rate < 70:
                        cell_b.fill = PatternFill(
                            start_color=_MISMATCH_FILL_HEX,
                            end_color=_MISMATCH_FILL_HEX,
                            fill_type="solid",
                        )
                except (ValueError, TypeError):
                    pass

            row_idx += 1

        # Only in A / B lists
        if summary.get("only_in_a"):
            row_idx += 1
            ws.cell(row=row_idx, column=1, value="Keys only in Source A:").font = Font(bold=True)
            row_idx += 1
            for key_val in summary["only_in_a"][:50]:
                ws.cell(row=row_idx, column=1, value=str(key_val))
                row_idx += 1

        if summary.get("only_in_b"):
            row_idx += 1
            ws.cell(row=row_idx, column=1, value="Keys only in Source B:").font = Font(bold=True)
            row_idx += 1
            for key_val in summary["only_in_b"][:50]:
                ws.cell(row=row_idx, column=1, value=str(key_val))
                row_idx += 1

        # Freeze panes below title
        ws.freeze_panes = "A5"

    @staticmethod
    def _write_table_sheet(
        ws: Any,
        rows: List[Dict[str, Any]],
        headers: List[str],
        row_keys: List[str],
        header_font: Any,
        header_fill: Any,
        header_alignment: Any,
        thin_border: Any,
        highlight_fill: Any = None,
    ) -> None:
        """Write a generic table sheet with headers and data rows."""
        if not rows:
            ws.cell(row=1, column=1, value="No data available for this section.")
            ws.cell(row=1, column=1).font = Font(italic=True, color="999999")
            return

        # Headers
        for col_idx, header_text in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, key in enumerate(row_keys, start=1):
                value = row_data.get(key, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                # Apply highlight if specified and applicable
                if highlight_fill is not None:
                    status = row_data.get("status", "")
                    if status == "mismatch":
                        cell.fill = highlight_fill

        # Freeze header
        ws.freeze_panes = "A2"

    @staticmethod
    def _write_unmatched_sheet(
        ws: Any,
        unmatched: Dict[str, List],
        header_font: Any,
        header_fill: Any,
        header_alignment: Any,
        thin_border: Any,
    ) -> None:
        """Write the Unmatched Records sheet."""
        only_a = unmatched.get("only_in_a", [])
        only_b = unmatched.get("only_in_b", [])

        row_idx = 1

        if not only_a and not only_b:
            ws.cell(row=row_idx, column=1, value="No unmatched records found.")
            ws.cell(row=row_idx, column=1).font = Font(italic=True, color="999999")
            return

        if only_a:
            ws.cell(row=row_idx, column=1, value="Keys Only in Source A").font = Font(bold=True, size=12)
            row_idx += 1
            cell = ws.cell(row=row_idx, column=1, value="Key")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            row_idx += 1

            for key_val in only_a:
                cell = ws.cell(row=row_idx, column=1, value=str(key_val))
                cell.border = thin_border
                cell.fill = PatternFill(
                    start_color=_MISMATCH_FILL_HEX,
                    end_color=_MISMATCH_FILL_HEX,
                    fill_type="solid",
                )
                row_idx += 1

            row_idx += 1

        if only_b:
            ws.cell(row=row_idx, column=1, value="Keys Only in Source B").font = Font(bold=True, size=12)
            row_idx += 1
            cell = ws.cell(row=row_idx, column=1, value="Key")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            row_idx += 1

            for key_val in only_b:
                cell = ws.cell(row=row_idx, column=1, value=str(key_val))
                cell.border = thin_border
                cell.fill = PatternFill(
                    start_color=_MISMATCH_FILL_HEX,
                    end_color=_MISMATCH_FILL_HEX,
                    fill_type="solid",
                )
                row_idx += 1

    @staticmethod
    def _write_log_sheet(
        ws: Any,
        log_entries: List[Dict[str, Any]],
        header_font: Any,
        header_fill: Any,
        header_alignment: Any,
        thin_border: Any,
    ) -> None:
        """Write a log sheet (filter log or normalization log)."""
        if not log_entries:
            ws.cell(row=1, column=1, value="No log entries for this section.")
            ws.cell(row=1, column=1).font = Font(italic=True, color="999999")
            return

        # Determine columns from the first entry
        first_entry = log_entries[0] if log_entries else {}
        # Standard log columns with sensible ordering
        preferred_keys = ["stage", "source", "action", "column", "rule", "status", "message", "rows_before", "rows_after", "rows_affected"]
        all_keys = list(first_entry.keys())
        # Order: preferred first, then remaining
        ordered_keys = [k for k in preferred_keys if k in all_keys]
        ordered_keys += [k for k in all_keys if k not in ordered_keys]

        # Headers
        for col_idx, key in enumerate(ordered_keys, start=1):
            header_text = key.replace("_", " ").title()
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        for row_idx, entry in enumerate(log_entries, start=2):
            for col_idx, key in enumerate(ordered_keys, start=1):
                value = entry.get(key, "")
                if isinstance(value, (list, dict)):
                    value = str(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                # Status-based colouring
                status = str(entry.get("status", ""))
                if status == "error":
                    cell.fill = PatternFill(
                        start_color=_MISMATCH_FILL_HEX,
                        end_color=_MISMATCH_FILL_HEX,
                        fill_type="solid",
                    )
                elif status == "warning":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        ws.freeze_panes = "A2"

    # ------------------------------------------------------------------
    # CSV generation
    # ------------------------------------------------------------------

    def _generate_csv(
        self,
        context: PipelineContext,
        config: Dict,
        output_dir: str,
        sections: List[str],
    ) -> None:
        """Generate one CSV file per enabled section.

        Parameters
        ----------
        context:
            Pipeline context with comparison results.
        config:
            Full pipeline configuration.
        output_dir:
            Directory where CSV files are written.
        sections:
            List of enabled section names.
        """
        os.makedirs(output_dir, exist_ok=True)

        for section_name in sections:
            content = self._apply_section_content(section_name, context, config)

            filename = f"{section_name}.csv"
            filepath = os.path.join(output_dir, filename)

            try:
                if section_name == "summary":
                    self._write_summary_csv(filepath, content)
                elif section_name == "mismatch_details":
                    self._write_table_csv(
                        filepath, content,
                        headers=["Key", "Column", "Value A", "Value B", "Diff Type"],
                        row_keys=["key", "column", "value_a", "value_b", "diff_type"],
                    )
                elif section_name == "unmatched_records":
                    self._write_unmatched_csv(filepath, content)
                elif section_name == "per_column_breakdown":
                    self._write_table_csv(
                        filepath, content,
                        headers=["Column", "Matched", "Mismatched", "Total", "Match Rate %"],
                        row_keys=["column", "matched", "mismatched", "total_compared", "match_rate_percent"],
                    )
                elif section_name in ("normalization_log", "filter_log"):
                    log_entries = (
                        context.normalization_log
                        if section_name == "normalization_log"
                        else context.filter_log
                    )
                    self._write_log_csv(filepath, log_entries)

            except Exception as exc:
                raise OutputError(
                    f"Failed to write CSV for section '{section_name}': {exc}",
                    {"rule": "RPT-002", "section": section_name},
                ) from exc

    @staticmethod
    def _write_summary_csv(filepath: str, summary: Dict[str, Any]) -> None:
        """Write summary data to CSV."""
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Metric", "Value"])
            writer.writerow(["Total Rows (Source A)", summary.get("total_rows_source_a", 0)])
            writer.writerow(["Total Rows (Source B)", summary.get("total_rows_source_b", 0)])
            writer.writerow(["Keys Compared", summary.get("total_compared_keys", 0)])
            writer.writerow(["Fully Matched Keys", summary.get("full_match_count", 0)])
            writer.writerow(["Keys with Mismatches", summary.get("partial_mismatch_count", 0)])
            writer.writerow(["Only in Source A", summary.get("only_in_a_count", 0)])
            writer.writerow(["Only in Source B", summary.get("only_in_b_count", 0)])
            writer.writerow(["Cell-Level Matches", summary.get("total_cell_matches", 0)])
            writer.writerow(["Cell-Level Mismatches", summary.get("total_cell_mismatches", 0)])
            writer.writerow(["Match Rate (%)", summary.get("match_rate_percent", 0)])

    @staticmethod
    def _write_table_csv(
        filepath: str,
        rows: List[Dict[str, Any]],
        headers: List[str],
        row_keys: List[str],
    ) -> None:
        """Write a generic table to CSV."""
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)

            if not rows:
                writer.writerow(["No data available"])
                return

            for row_data in rows:
                writer.writerow([row_data.get(key, "") for key in row_keys])

    @staticmethod
    def _write_unmatched_csv(filepath: str, unmatched: Dict[str, List]) -> None:
        """Write unmatched records to CSV."""
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Source", "Key"])

            only_a = unmatched.get("only_in_a", [])
            only_b = unmatched.get("only_in_b", [])

            if not only_a and not only_b:
                writer.writerow(["No unmatched records found"])
                return

            for key_val in only_a:
                writer.writerow(["Source A", str(key_val)])
            for key_val in only_b:
                writer.writerow(["Source B", str(key_val)])

    @staticmethod
    def _write_log_csv(filepath: str, log_entries: List[Dict[str, Any]]) -> None:
        """Write log entries to CSV."""
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)

            if not log_entries:
                writer.writerow(["No log entries"])
                return

            first_entry = log_entries[0]
            preferred_keys = ["stage", "source", "action", "column", "rule", "status", "message", "rows_before", "rows_after", "rows_affected"]
            all_keys = list(first_entry.keys())
            ordered_keys = [k for k in preferred_keys if k in all_keys]
            ordered_keys += [k for k in all_keys if k not in ordered_keys]

            writer.writerow(ordered_keys)
            for entry in log_entries:
                row = []
                for key in ordered_keys:
                    value = entry.get(key, "")
                    if isinstance(value, (list, dict)):
                        value = str(value)
                    row.append(value)
                writer.writerow(row)

    # ------------------------------------------------------------------
    # HTML generation
    # ------------------------------------------------------------------

    def _generate_html(
        self,
        context: PipelineContext,
        config: Dict,
        output_path: str,
        sections: List[str],
        title: str,
    ) -> None:
        """Generate a self-contained HTML report with inline CSS.

        Features collapsible sections, styled tables, progress bars
        for match rates, and a clean professional design (RPT-005).

        Parameters
        ----------
        context:
            Pipeline context with comparison results.
        config:
            Full pipeline configuration.
        output_path:
            File path for the HTML output.
        sections:
            List of enabled section names.
        title:
            Report title.
        """
        stats = context.stats or {}
        summary = self._get_summary_data(context, stats)

        # Build section HTML
        section_html_parts: List[str] = []

        for section_name in sections:
            content = self._apply_section_content(section_name, context, config)
            section_html = self._build_section_html(section_name, content, context)
            section_html_parts.append(section_html)

        sections_html = "\n".join(section_html_parts)

        now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{_html_escape(title)}</title>
    <style>
        :root {{
            --primary: #2563eb;
            --primary-dark: #1d4ed8;
            --success: #16a34a;
            --danger: #dc2626;
            --warning: #d97706;
            --gray-50: #f9fafb;
            --gray-100: #f3f4f6;
            --gray-200: #e5e7eb;
            --gray-300: #d1d5db;
            --gray-500: #6b7280;
            --gray-700: #374151;
            --gray-900: #111827;
            --white: #ffffff;
            --shadow: 0 1px 3px 0 rgba(0,0,0,0.1), 0 1px 2px -1px rgba(0,0,0,0.1);
            --shadow-md: 0 4px 6px -1px rgba(0,0,0,0.1), 0 2px 4px -2px rgba(0,0,0,0.1);
        }}

        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
            background-color: var(--gray-50);
            color: var(--gray-900);
            line-height: 1.6;
        }}

        .container {{
            max-width: 1200px;
            margin: 0 auto;
            padding: 24px 16px;
        }}

        .report-header {{
            background: linear-gradient(135deg, var(--primary), var(--primary-dark));
            color: var(--white);
            padding: 32px 40px;
            border-radius: 12px;
            margin-bottom: 24px;
            box-shadow: var(--shadow-md);
        }}

        .report-header h1 {{
            font-size: 28px;
            font-weight: 700;
            margin-bottom: 8px;
        }}

        .report-header .timestamp {{
            font-size: 14px;
            opacity: 0.85;
        }}

        .summary-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(200px, 1fr));
            gap: 16px;
            margin-bottom: 24px;
        }}

        .metric-card {{
            background: var(--white);
            border-radius: 10px;
            padding: 20px;
            box-shadow: var(--shadow);
            border-left: 4px solid var(--primary);
        }}

        .metric-card.success {{ border-left-color: var(--success); }}
        .metric-card.danger {{ border-left-color: var(--danger); }}
        .metric-card.warning {{ border-left-color: var(--warning); }}

        .metric-card .label {{
            font-size: 12px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            color: var(--gray-500);
            margin-bottom: 4px;
        }}

        .metric-card .value {{
            font-size: 28px;
            font-weight: 700;
            color: var(--gray-900);
        }}

        .metric-card .value.success {{ color: var(--success); }}
        .metric-card .value.danger {{ color: var(--danger); }}

        .section {{
            background: var(--white);
            border-radius: 10px;
            box-shadow: var(--shadow);
            margin-bottom: 16px;
            overflow: hidden;
        }}

        .section-header {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            padding: 16px 24px;
            cursor: pointer;
            user-select: none;
            transition: background-color 0.15s;
        }}

        .section-header:hover {{
            background-color: var(--gray-100);
        }}

        .section-header h2 {{
            font-size: 16px;
            font-weight: 600;
            color: var(--gray-900);
        }}

        .section-toggle {{
            width: 24px;
            height: 24px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 18px;
            color: var(--gray-500);
            transition: transform 0.2s;
        }}

        .section.collapsed .section-toggle {{
            transform: rotate(-90deg);
        }}

        .section-content {{
            padding: 0 24px 24px 24px;
        }}

        .section.collapsed .section-content {{
            display: none;
        }}

        .progress-bar-container {{
            width: 100%;
            background-color: var(--gray-200);
            border-radius: 8px;
            height: 24px;
            overflow: hidden;
            margin-top: 8px;
        }}

        .progress-bar {{
            height: 100%;
            border-radius: 8px;
            transition: width 0.3s ease;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 12px;
            font-weight: 600;
            color: var(--white);
        }}

        .progress-bar.high {{ background-color: var(--success); }}
        .progress-bar.medium {{ background-color: var(--warning); }}
        .progress-bar.low {{ background-color: var(--danger); }}

        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 14px;
        }}

        table thead th {{
            background-color: var(--gray-100);
            padding: 10px 12px;
            text-align: left;
            font-weight: 600;
            color: var(--gray-700);
            border-bottom: 2px solid var(--gray-200);
            font-size: 12px;
            text-transform: uppercase;
            letter-spacing: 0.3px;
            position: sticky;
            top: 0;
        }}

        table tbody td {{
            padding: 8px 12px;
            border-bottom: 1px solid var(--gray-200);
            color: var(--gray-700);
        }}

        table tbody tr:hover {{
            background-color: var(--gray-50);
        }}

        table tbody tr.match-row td {{
            background-color: #dcfce7;
        }}

        table tbody tr.mismatch-row td {{
            background-color: #fee2e2;
        }}

        .table-wrapper {{
            max-height: 400px;
            overflow-y: auto;
            border: 1px solid var(--gray-200);
            border-radius: 8px;
        }}

        .badge {{
            display: inline-block;
            padding: 2px 8px;
            border-radius: 12px;
            font-size: 11px;
            font-weight: 600;
        }}

        .badge-success {{
            background-color: #dcfce7;
            color: var(--success);
        }}

        .badge-danger {{
            background-color: #fee2e2;
            color: var(--danger);
        }}

        .badge-warning {{
            background-color: #fef3c7;
            color: var(--warning);
        }}

        .badge-info {{
            background-color: #dbeafe;
            color: var(--primary);
        }}

        .empty-state {{
            text-align: center;
            padding: 32px;
            color: var(--gray-500);
            font-style: italic;
        }}

        .two-col-list {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 16px;
        }}

        .key-list {{
            list-style: none;
            padding: 0;
        }}

        .key-list li {{
            padding: 6px 10px;
            border-radius: 4px;
            font-size: 13px;
            font-family: 'SF Mono', 'Fira Code', monospace;
            background-color: var(--gray-50);
            margin-bottom: 4px;
        }}

        @media (max-width: 768px) {{
            .summary-grid {{
                grid-template-columns: repeat(2, 1fr);
            }}
            .two-col-list {{
                grid-template-columns: 1fr;
            }}
            .report-header {{
                padding: 24px;
            }}
            .report-header h1 {{
                font-size: 22px;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="report-header">
            <h1>{_html_escape(title)}</h1>
            <div class="timestamp">Generated: {now_str}</div>
        </div>

        <div class="summary-grid">
            <div class="metric-card">
                <div class="label">Match Rate</div>
                <div class="value {'success' if summary.get('match_rate_percent', 0) >= 90 else 'danger' if summary.get('match_rate_percent', 0) < 70 else ''}">
                    {summary.get('match_rate_percent', 0)}%
                </div>
                <div class="progress-bar-container">
                    <div class="progress-bar {'high' if summary.get('match_rate_percent', 0) >= 90 else 'medium' if summary.get('match_rate_percent', 0) >= 70 else 'low'}" style="width: {min(summary.get('match_rate_percent', 0), 100)}%">
                        {summary.get('match_rate_percent', 0)}%
                    </div>
                </div>
            </div>
            <div class="metric-card success">
                <div class="label">Full Matches</div>
                <div class="value success">{summary.get('full_match_count', 0)}</div>
            </div>
            <div class="metric-card danger">
                <div class="label">Mismatches</div>
                <div class="value danger">{summary.get('total_cell_mismatches', 0)}</div>
            </div>
            <div class="metric-card">
                <div class="label">Only in A</div>
                <div class="value">{summary.get('only_in_a_count', 0)}</div>
            </div>
            <div class="metric-card">
                <div class="label">Only in B</div>
                <div class="value">{summary.get('only_in_b_count', 0)}</div>
            </div>
            <div class="metric-card">
                <div class="label">Keys Compared</div>
                <div class="value">{summary.get('total_compared_keys', 0)}</div>
            </div>
        </div>

        {sections_html}
    </div>

    <script>
        document.querySelectorAll('.section-header').forEach(function(header) {{
            header.addEventListener('click', function() {{
                this.parentElement.classList.toggle('collapsed');
            }});
        }});
    </script>
</body>
</html>"""

        try:
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(html)
        except Exception as exc:
            raise OutputError(
                f"Failed to write HTML file: {output_path}",
                {"rule": "RPT-002", "path": output_path, "error": str(exc)},
            ) from exc

    def _build_section_html(
        self,
        section_name: str,
        content: Any,
        context: PipelineContext,
    ) -> str:
        """Build the HTML for a single section."""
        display_name = section_name.replace("_", " ").title()

        if section_name == "mismatch_details":
            return self._build_mismatch_html(display_name, content)
        elif section_name == "unmatched_records":
            return self._build_unmatched_html(display_name, content)
        elif section_name == "per_column_breakdown":
            return self._build_column_breakdown_html(display_name, content)
        elif section_name == "normalization_log":
            return self._build_log_html(display_name, context.normalization_log)
        elif section_name == "filter_log":
            return self._build_log_html(display_name, context.filter_log)
        else:
            return self._build_generic_section_html(display_name, content)

    @staticmethod
    def _build_mismatch_html(display_name: str, mismatches: List[Dict]) -> str:
        """Build HTML for mismatch details section."""
        if not mismatches:
            return f"""
            <div class="section">
                <div class="section-header">
                    <h2>{display_name}</h2>
                    <div class="section-toggle">&#9660;</div>
                </div>
                <div class="section-content">
                    <div class="empty-state">No mismatches found. All compared values match.</div>
                </div>
            </div>"""

        rows_html = ""
        for m in mismatches[:500]:  # Limit for performance
            diff_type = m.get("diff_type", "value_diff")
            badge_class = "badge-danger"
            badge_text = diff_type.replace("_", " ").title()

            rows_html += f"""
                <tr class="mismatch-row">
                    <td>{_html_escape(str(m.get('key', '')))}</td>
                    <td>{_html_escape(str(m.get('column', '')))}</td>
                    <td><code>{_html_escape(str(m.get('value_a', '')))}</code></td>
                    <td><code>{_html_escape(str(m.get('value_b', '')))}</code></td>
                    <td><span class="badge {badge_class}">{_html_escape(badge_text)}</span></td>
                </tr>"""

        overflow_note = ""
        if len(mismatches) > 500:
            overflow_note = f'<div class="empty-state">Showing first 500 of {len(mismatches)} mismatches.</div>'

        return f"""
        <div class="section">
            <div class="section-header">
                <h2>{display_name} <span class="badge badge-danger">{len(mismatches)}</span></h2>
                <div class="section-toggle">&#9660;</div>
            </div>
            <div class="section-content">
                <div class="table-wrapper">
                    <table>
                        <thead>
                            <tr>
                                <th>Key</th>
                                <th>Column</th>
                                <th>Value A</th>
                                <th>Value B</th>
                                <th>Diff Type</th>
                            </tr>
                        </thead>
                        <tbody>
                            {rows_html}
                        </tbody>
                    </table>
                </div>
                {overflow_note}
            </div>
        </div>"""

    @staticmethod
    def _build_unmatched_html(display_name: str, unmatched: Dict[str, List]) -> str:
        """Build HTML for unmatched records section."""
        only_a = unmatched.get("only_in_a", [])
        only_b = unmatched.get("only_in_b", [])

        if not only_a and not only_b:
            return f"""
            <div class="section">
                <div class="section-header">
                    <h2>{display_name}</h2>
                    <div class="section-toggle">&#9660;</div>
                </div>
                <div class="section-content">
                    <div class="empty-state">No unmatched records found.</div>
                </div>
            </div>"""

        a_list_html = ""
        for key_val in only_a[:100]:
            a_list_html += f"<li>{_html_escape(str(key_val))}</li>"
        if len(only_a) > 100:
            a_list_html += f"<li>... and {len(only_a) - 100} more</li>"

        b_list_html = ""
        for key_val in only_b[:100]:
            b_list_html += f"<li>{_html_escape(str(key_val))}</li>"
        if len(only_b) > 100:
            b_list_html += f"<li>... and {len(only_b) - 100} more</li>"

        return f"""
        <div class="section">
            <div class="section-header">
                <h2>{display_name}</h2>
                <div class="section-toggle">&#9660;</div>
            </div>
            <div class="section-content">
                <div class="two-col-list">
                    <div>
                        <h3 style="margin-bottom: 8px; color: var(--gray-700); font-size: 14px;">
                            Only in Source A <span class="badge badge-danger">{len(only_a)}</span>
                        </h3>
                        <ul class="key-list">
                            {a_list_html if a_list_html else '<li style="color: var(--gray-500); font-style: italic;">None</li>'}
                        </ul>
                    </div>
                    <div>
                        <h3 style="margin-bottom: 8px; color: var(--gray-700); font-size: 14px;">
                            Only in Source B <span class="badge badge-danger">{len(only_b)}</span>
                        </h3>
                        <ul class="key-list">
                            {b_list_html if b_list_html else '<li style="color: var(--gray-500); font-style: italic;">None</li>'}
                        </ul>
                    </div>
                </div>
            </div>
        </div>"""

    @staticmethod
    def _build_column_breakdown_html(display_name: str, columns: List[Dict]) -> str:
        """Build HTML for per-column breakdown section."""
        if not columns:
            return f"""
            <div class="section">
                <div class="section-header">
                    <h2>{display_name}</h2>
                    <div class="section-toggle">&#9660;</div>
                </div>
                <div class="section-content">
                    <div class="empty-state">No column comparison data available.</div>
                </div>
            </div>"""

        rows_html = ""
        for col in columns:
            rate = col.get("match_rate_percent", 0)
            bar_class = "high" if rate >= 90 else "medium" if rate >= 70 else "low"
            badge_class = "badge-success" if rate >= 90 else "badge-warning" if rate >= 70 else "badge-danger"

            rows_html += f"""
                <tr>
                    <td><code>{_html_escape(str(col.get('column', '')))}</code></td>
                    <td>{col.get('matched', 0)}</td>
                    <td>{col.get('mismatched', 0)}</td>
                    <td>{col.get('total_compared', 0)}</td>
                    <td>
                        <span class="badge {badge_class}">{rate}%</span>
                        <div class="progress-bar-container" style="width: 120px; display: inline-block; vertical-align: middle; margin-left: 8px;">
                            <div class="progress-bar {bar_class}" style="width: {min(rate, 100)}%; font-size: 10px;">
                                {rate}%
                            </div>
                        </div>
                    </td>
                </tr>"""

        return f"""
        <div class="section">
            <div class="section-header">
                <h2>{display_name}</h2>
                <div class="section-toggle">&#9660;</div>
            </div>
            <div class="section-content">
                <div class="table-wrapper">
                    <table>
                        <thead>
                            <tr>
                                <th>Column</th>
                                <th>Matched</th>
                                <th>Mismatched</th>
                                <th>Total</th>
                                <th>Match Rate</th>
                            </tr>
                        </thead>
                        <tbody>
                            {rows_html}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>"""

    @staticmethod
    def _build_log_html(display_name: str, log_entries: List[Dict]) -> str:
        """Build HTML for a log section (filter or normalization log)."""
        if not log_entries:
            return f"""
            <div class="section">
                <div class="section-header">
                    <h2>{display_name}</h2>
                    <div class="section-toggle">&#9660;</div>
                </div>
                <div class="section-content">
                    <div class="empty-state">No log entries for this section.</div>
                </div>
            </div>"""

        preferred_keys = ["stage", "source", "action", "column", "status", "message", "rows_before", "rows_after", "rows_affected"]
        first_entry = log_entries[0] if log_entries else {}
        all_keys = list(first_entry.keys())
        ordered_keys = [k for k in preferred_keys if k in all_keys]
        ordered_keys += [k for k in all_keys if k not in ordered_keys]

        # Only show the most useful columns
        display_keys = [k for k in ordered_keys if k in ("stage", "source", "action", "status", "message")]

        headers_html = "".join(
            f"<th>{k.replace('_', ' ').title()}</th>" for k in display_keys
        )

        rows_html = ""
        for entry in log_entries[:200]:
            status = str(entry.get("status", ""))
            badge_class = "badge-danger" if status == "error" else "badge-warning" if status == "warning" else "badge-success"

            cells_html = ""
            for key in display_keys:
                value = entry.get(key, "")
                if isinstance(value, (list, dict)):
                    value = str(value)
                if key == "status":
                    cells_html += f'<td><span class="badge {badge_class}">{_html_escape(str(value))}</span></td>'
                else:
                    cells_html += f"<td>{_html_escape(str(value))}</td>"

            rows_html += f"<tr>{cells_html}</tr>"

        overflow_note = ""
        if len(log_entries) > 200:
            overflow_note = f'<div class="empty-state">Showing first 200 of {len(log_entries)} entries.</div>'

        return f"""
        <div class="section">
            <div class="section-header">
                <h2>{display_name} <span class="badge badge-info">{len(log_entries)}</span></h2>
                <div class="section-toggle">&#9660;</div>
            </div>
            <div class="section-content">
                <div class="table-wrapper">
                    <table>
                        <thead>
                            <tr>{headers_html}</tr>
                        </thead>
                        <tbody>
                            {rows_html}
                        </tbody>
                    </table>
                </div>
                {overflow_note}
            </div>
        </div>"""

    @staticmethod
    def _build_generic_section_html(display_name: str, content: Any) -> str:
        """Build HTML for a generic (summary or unknown) section."""
        return f"""
        <div class="section">
            <div class="section-header">
                <h2>{display_name}</h2>
                <div class="section-toggle">&#9660;</div>
            </div>
            <div class="section-content">
                <div class="empty-state">Section content rendered in summary cards above.</div>
            </div>
        </div>"""

    # ------------------------------------------------------------------
    # Markdown generation
    # ------------------------------------------------------------------

    def _generate_markdown(
        self,
        context: PipelineContext,
        config: Dict,
        output_path: str,
        sections: List[str],
        title: str,
    ) -> None:
        """Generate a Markdown report.

        Parameters
        ----------
        context:
            Pipeline context with comparison results.
        config:
            Full pipeline configuration.
        output_path:
            File path for the Markdown output.
        sections:
            List of enabled section names.
        title:
            Report title.
        """
        stats = context.stats or {}
        summary = self._get_summary_data(context, stats)
        now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

        md_parts: List[str] = []

        md_parts.append(f"# {title}")
        md_parts.append(f"\n> Generated: {now_str}\n")

        # Summary (always first)
        md_parts.append("## Summary\n")
        md_parts.append("| Metric | Value |")
        md_parts.append("|--------|-------|")
        md_parts.append(f"| Total Rows (Source A) | {summary.get('total_rows_source_a', 0)} |")
        md_parts.append(f"| Total Rows (Source B) | {summary.get('total_rows_source_b', 0)} |")
        md_parts.append(f"| Keys Compared | {summary.get('total_compared_keys', 0)} |")
        md_parts.append(f"| Fully Matched Keys | {summary.get('full_match_count', 0)} |")
        md_parts.append(f"| Keys with Mismatches | {summary.get('partial_mismatch_count', 0)} |")
        md_parts.append(f"| Only in Source A | {summary.get('only_in_a_count', 0)} |")
        md_parts.append(f"| Only in Source B | {summary.get('only_in_b_count', 0)} |")
        md_parts.append(f"| Cell-Level Matches | {summary.get('total_cell_matches', 0)} |")
        md_parts.append(f"| Cell-Level Mismatches | {summary.get('total_cell_mismatches', 0)} |")
        md_parts.append(f"| **Match Rate** | **{summary.get('match_rate_percent', 0)}%** |")
        md_parts.append("")

        for section_name in sections:
            if section_name == "summary":
                continue  # Already rendered

            content = self._apply_section_content(section_name, context, config)
            md_section = self._build_section_markdown(section_name, content, context)
            md_parts.append(md_section)

        markdown_text = "\n".join(md_parts)

        try:
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(markdown_text)
        except Exception as exc:
            raise OutputError(
                f"Failed to write Markdown file: {output_path}",
                {"rule": "RPT-002", "path": output_path, "error": str(exc)},
            ) from exc

    @staticmethod
    def _build_section_markdown(
        section_name: str,
        content: Any,
        context: PipelineContext,
    ) -> str:
        """Build Markdown for a single section."""
        display_name = section_name.replace("_", " ").title()
        parts: List[str] = []

        if section_name == "mismatch_details":
            mismatches = content if isinstance(content, list) else []
            parts.append(f"## {display_name}\n")

            if not mismatches:
                parts.append("*No mismatches found. All compared values match.*\n")
            else:
                parts.append("| Key | Column | Value A | Value B | Diff Type |")
                parts.append("|-----|--------|---------|---------|-----------|")
                for m in mismatches[:200]:
                    parts.append(
                        f"| {m.get('key', '')} | {m.get('column', '')} | "
                        f"{m.get('value_a', '')} | {m.get('value_b', '')} | "
                        f"{m.get('diff_type', '')} |"
                    )
                if len(mismatches) > 200:
                    parts.append(f"\n*Showing first 200 of {len(mismatches)} mismatches.*")
            parts.append("")

        elif section_name == "unmatched_records":
            unmatched = content if isinstance(content, dict) else {}
            only_a = unmatched.get("only_in_a", [])
            only_b = unmatched.get("only_in_b", [])
            parts.append(f"## {display_name}\n")

            if not only_a and not only_b:
                parts.append("*No unmatched records found.*\n")
            else:
                parts.append(f"### Only in Source A ({len(only_a)})\n")
                if only_a:
                    for key_val in only_a[:50]:
                        parts.append(f"- `{key_val}`")
                    if len(only_a) > 50:
                        parts.append(f"- ... and {len(only_a) - 50} more")
                else:
                    parts.append("*None*")
                parts.append("")

                parts.append(f"### Only in Source B ({len(only_b)})\n")
                if only_b:
                    for key_val in only_b[:50]:
                        parts.append(f"- `{key_val}`")
                    if len(only_b) > 50:
                        parts.append(f"- ... and {len(only_b) - 50} more")
                else:
                    parts.append("*None*")
                parts.append("")

        elif section_name == "per_column_breakdown":
            columns = content if isinstance(content, list) else []
            parts.append(f"## {display_name}\n")

            if not columns:
                parts.append("*No column comparison data available.*\n")
            else:
                parts.append("| Column | Matched | Mismatched | Total | Match Rate |")
                parts.append("|--------|---------|------------|-------|------------|")
                for col in columns:
                    rate = col.get("match_rate_percent", 0)
                    parts.append(
                        f"| {col.get('column', '')} | {col.get('matched', 0)} | "
                        f"{col.get('mismatched', 0)} | {col.get('total_compared', 0)} | "
                        f"{rate}% |"
                    )
                parts.append("")

        elif section_name in ("normalization_log", "filter_log"):
            log_entries = (
                context.normalization_log
                if section_name == "normalization_log"
                else context.filter_log
            )
            parts.append(f"## {display_name}\n")

            if not log_entries:
                parts.append("*No log entries for this section.*\n")
            else:
                display_keys = ["stage", "source", "action", "status", "message"]
                first_entry = log_entries[0] if log_entries else {}
                available_keys = [k for k in display_keys if k in first_entry]

                if available_keys:
                    header = " | ".join(k.replace("_", " ").title() for k in available_keys)
                    separator = " | ".join("---" for _ in available_keys)
                    parts.append(f"| {header} |")
                    parts.append(f"| {separator} |")
                    for entry in log_entries[:100]:
                        row = " | ".join(
                            str(entry.get(k, "")).replace("|", "\\|") for k in available_keys
                        )
                        parts.append(f"| {row} |")
                    if len(log_entries) > 100:
                        parts.append(f"\n*Showing first 100 of {len(log_entries)} entries.*")
                parts.append("")

        return "\n".join(parts)

    # ------------------------------------------------------------------
    # Section content router
    # ------------------------------------------------------------------

    def _apply_section_content(
        self,
        section_name: str,
        context: PipelineContext,
        config: Dict,
    ) -> Any:
        """Route to section-specific content generators.

        Returns the appropriate data structure for each section type.
        """
        stats = context.stats or {}

        if section_name == "summary":
            return self._get_summary_data(context, stats)
        elif section_name == "mismatch_details":
            return self._get_mismatch_details(context)
        elif section_name == "unmatched_records":
            return self._get_unmatched_records(context)
        elif section_name == "per_column_breakdown":
            return self._get_per_column_breakdown(context)
        elif section_name == "normalization_log":
            return context.normalization_log
        elif section_name == "filter_log":
            return context.filter_log
        else:
            return None

    # ------------------------------------------------------------------
    # Utilities
    # ------------------------------------------------------------------

    # _html_escape is defined at module level for use in both instance
    # and static method f-strings.
